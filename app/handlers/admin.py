import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, BufferedInputFile, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.filters import Command
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from sqlalchemy import select, and_
from ..database import async_session_maker
from ..models import Trip, Booking
from ..keyboards import (
    admin_main_menu, admin_trips_list_raw, admin_trip_menu,
    admin_bookings_filter, admin_edit_trip, back_button, bus_type_kb,
    broadcast_confirm_kb, main_menu,
)
from ..config import ADMIN_IDS, ADMIN_PASSWORD
from ..utils import (
    format_date_ru, format_price, format_price_verbose,
    now_utc, get_occupied_seats, get_trip_stats, get_trip_total_seats,
    render_seat_map, get_occupied_seat_ids, STATUS_EMOJI,
)
from .booking import payment_method_label

router = Router()

admin_sessions = set()


class AdminStates(StatesGroup):
    waiting_password = State()
    create_date = State()
    create_bus_type = State()
    create_seats_left = State()
    create_seats_middle = State()
    create_seats_right = State()
    create_price_adult = State()
    create_price_child = State()
    create_location = State()
    create_time = State()
    create_bus = State()
    edit_value = State()
    broadcast_text = State()


def is_admin(telegram_id: int) -> bool:
    return telegram_id in ADMIN_IDS or telegram_id in admin_sessions


@router.message(Command("admin"))
async def cmd_admin(message: Message, state: FSMContext):
    if message.from_user.id in ADMIN_IDS:
        admin_sessions.add(message.from_user.id)
        await message.answer("🔐 Панель администратора:", reply_markup=admin_main_menu())
        return

    await state.set_state(AdminStates.waiting_password)
    await message.answer("🔐 Введите пароль для доступа к админ-панели:")


@router.message(AdminStates.waiting_password)
async def process_admin_password(message: Message, state: FSMContext):
    if message.text and message.text.strip() == ADMIN_PASSWORD:
        admin_sessions.add(message.from_user.id)
        await state.clear()
        await message.answer("🔐 Панель администратора:", reply_markup=admin_main_menu())
    else:
        await message.answer("❌ Неверный пароль\nПопробуйте снова или отправьте /start")


@router.callback_query(F.data == "admin_menu")
async def admin_menu_handler(callback: CallbackQuery):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        await callback.message.edit_text("🚫 Доступ запрещён")
        return
    await callback.message.edit_text("🔐 Панель администратора:", reply_markup=admin_main_menu())


@router.callback_query(F.data == "admin_trips")
async def admin_list_trips(callback: CallbackQuery):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    async with async_session_maker() as session:
        result = await session.execute(
            select(Trip).order_by(Trip.date.desc())
        )
        trips = result.scalars().all()

        trip_list = []
        for trip in trips:
            occ = await get_occupied_seats(session, trip.id)
            total = get_trip_total_seats(trip)
            trip_list.append({
                "id": trip.id,
                "date": trip.date,
                "max_seats": total,
                "occupied": occ,
                "bus_type": trip.bus_type,
            })

    if not trip_list:
        await callback.message.edit_text(
            "📋 Нет созданных рейсов",
            reply_markup=back_button("admin_menu"),
        )
        return

    await callback.message.edit_text(
        "📋 Список рейсов:",
        reply_markup=admin_trips_list_raw(trip_list),
    )


@router.callback_query(F.data.startswith("admin_trip_"))
async def admin_trip_detail(callback: CallbackQuery):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    trip_id = int(callback.data.split("_")[-1])

    async with async_session_maker() as session:
        result = await session.execute(select(Trip).where(Trip.id == trip_id))
        trip = result.scalar_one_or_none()

        if not trip:
            await callback.message.edit_text("🚌 Рейс не найден", reply_markup=back_button("admin_trips"))
            return

        occupied = await get_occupied_seats(session, trip_id)
        total_seats = get_trip_total_seats(trip)
        occupied_ids = await get_occupied_seat_ids(session, trip_id)
        seat_map = render_seat_map(
            trip.bus_type, trip.seats_left, trip.seats_middle,
            trip.seats_right, occupied_ids,
        )

        booking_result = await session.execute(
            select(Booking).where(Booking.trip_id == trip_id).order_by(Booking.created_at)
        )
        bookings = booking_result.scalars().all()

        trip_data = {
            "id": trip.id,
            "date_str": format_date_ru(trip.date),
            "max_seats": total_seats,
            "occupied": occupied,
            "price_adult_str": format_price(trip.price_adult),
            "price_child_str": format_price(trip.price_child),
            "pickup_location": trip.pickup_location,
            "departure_time": trip.departure_time or "—",
            "bus_number": trip.bus_number or "—",
            "bus_type": "VIP" if trip.bus_type == "vip" else "Sleeper",
            "is_active": trip.is_active,
        }

        booking_lines = []
        for b in bookings:
            username = b.user.username or str(b.user.telegram_id)
            status = STATUS_EMOJI.get(b.status, b.status)
            total_people = b.adults + b.children
            seats_str = f" [{b.selected_seats}]" if b.selected_seats else ""
            line = (
                f"  @{username} | {b.adults}взр+{b.children}дет={total_people}чел{seats_str} | "
                f"{format_price_verbose(b.total_price, b.payment_method)} | {status}"
            )
            if b.pending_add_people:
                line += f" | ⏳ +{b.pending_add_people}чел без оплаты"
            if b.pending_add_amount:
                line += f" | ⏳ доплата {format_price(b.pending_add_amount)}"
            booking_lines.append(line)
            for p in b.payments:
                mark = "✅" if p.status == "paid" else "⏳"
                hist = f"    💳 {mark} {payment_method_label(p.method)} {format_price(p.amount)}"
                if (p.comment or "").strip():
                    hist += f" · {p.comment.strip()}"
                booking_lines.append(hist)

    active_status = "✅ активен" if trip_data["is_active"] else "❌ неактивен"

    text = (
        f"🚌 Рейс: {trip_data['date_str']}\n"
        f"⏰ Отправление: {trip_data['departure_time']}\n"
        f"🚌 Тип: {trip_data['bus_type']}\n"
        f"💺 Мест: {trip_data['occupied']}/{trip_data['max_seats']}\n"
        f"💰 Цена взр.: {trip_data['price_adult_str']}\n"
        f"💰 Цена дет.: {trip_data['price_child_str']}\n"
        f"📍 Сбор: {trip_data['pickup_location']}\n"
        f"🚌 Номер: {trip_data['bus_number']}\n"
        f"📊 Статус: {active_status}\n\n"
        f"{seat_map}\n"
    )

    if booking_lines:
        text += "📋 Записи:\n" + "\n".join(booking_lines)
    else:
        text += "Нет записей"

    await callback.message.edit_text(text, reply_markup=admin_trip_menu(trip_id))


@router.callback_query(F.data.startswith("admin_bookings_"))
async def admin_show_bookings(callback: CallbackQuery):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    parts = callback.data.split("_")
    trip_id = int(parts[2])
    status_filter = parts[3] if len(parts) > 3 else None

    if status_filter is None:
        await callback.message.edit_text(
            "❓ Выберите статус для фильтрации:",
            reply_markup=admin_bookings_filter(trip_id),
        )
        return

    async with async_session_maker() as session:
        query = select(Booking).where(Booking.trip_id == trip_id)
        if status_filter != "all":
            query = query.where(Booking.status == status_filter)
        query = query.order_by(Booking.created_at)
        result = await session.execute(query)
        bookings = result.scalars().all()

        trip_result = await session.execute(select(Trip).where(Trip.id == trip_id))
        trip = trip_result.scalar_one_or_none()

        trip_date_str = format_date_ru(trip.date) if trip else "?"

        booking_list = []
        for b in bookings:
            status_label = STATUS_EMOJI.get(b.status, b.status)
            if b.pending_add_people:
                status_label += f" ⏳+{b.pending_add_people}чел без оплаты"
            if b.pending_add_amount:
                status_label += f" ⏳+{format_price(b.pending_add_amount)}"
            booking_list.append({
                "id": b.id,
                "username": b.user.username or str(b.user.telegram_id),
                "telegram_id": b.user.telegram_id,
                "adults": b.adults,
                "children": b.children,
                "total_people": b.adults + b.children,
                "total_price": format_price_verbose(b.total_price, b.payment_method),
                "status": status_label,
            })

    if not booking_list:
        await callback.message.edit_text(
            "📋 Нет броней с выбранным статусом",
            reply_markup=back_button(f"admin_bookings_{trip_id}"),
        )
        return

    status_name = {"all": "📋 все", "pending": "⏳ ожидают оплаты", "paid": "✅ оплачено"}
    lines = [f"📋 Брони на {trip_date_str}", f"Фильтр: {status_name.get(status_filter, status_filter)}\n"]

    for b in booking_list:
        lines.append(
            f"#{b['id']} @{b['username']} | {b['adults']}взр+{b['children']}дет={b['total_people']}чел | "
            f"{b['total_price']} | {b['status']}"
        )

    await callback.message.edit_text(
        "\n".join(lines),
        reply_markup=back_button(f"admin_bookings_{trip_id}"),
    )


@router.callback_query(F.data.startswith("booking_paid_"))
async def admin_booking_paid(callback: CallbackQuery):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    booking_id = int(callback.data.split("_")[-1])

    async with async_session_maker() as session:
        result = await session.execute(select(Booking).where(Booking.id == booking_id))
        booking = result.scalar_one_or_none()

        if not booking:
            await callback.message.edit_text("📋 Бронь не найдена")
            return

        booking.status = "paid"
        booking.pending_add_amount = None
        booking.pending_add_people = None
        for p in booking.payments:
            if p.status != "paid":
                p.status = "paid"
                p.paid_at = now_utc()
        await session.commit()

        trip = booking.trip
        user = booking.user

    await callback.message.edit_text(
        f"✅ Бронь #{booking_id} отмечена как оплачена",
        reply_markup=back_button("admin_trips"),
    )

    try:
        await callback.bot.send_message(
            user.telegram_id,
            f"✅ Ваша бронь на {format_date_ru(trip.date)} оплачена!\n"
            f"Спасибо, ждём вас в рейс!",
        )
    except Exception:
        pass

    for admin_id in ADMIN_IDS:
        if admin_id == callback.from_user.id:
            continue
        try:
            await callback.bot.send_message(
                admin_id,
                f"✅ Бронь #{booking_id} подтверждена как оплаченная\n"
                f"👤 Пользователь: @{user.username or user.telegram_id}\n"
                f"🚌 Рейс: {format_date_ru(trip.date)}\n"
                f"💰 Сумма: {format_price_verbose(booking.total_price, booking.payment_method)}",
            )
        except Exception:
            pass


@router.callback_query(F.data.startswith("booking_cancel_"))
async def admin_booking_cancel(callback: CallbackQuery):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    booking_id = int(callback.data.split("_")[-1])

    async with async_session_maker() as session:
        result = await session.execute(select(Booking).where(Booking.id == booking_id))
        booking = result.scalar_one_or_none()

        if not booking:
            await callback.message.edit_text("📋 Бронь не найдена")
            return

        booking.status = "cancelled"
        await session.commit()

        trip = booking.trip
        user = booking.user

    await callback.message.edit_text(
        f"❌ Бронь #{booking_id} отменена",
        reply_markup=back_button("admin_trips"),
    )

    try:
        await callback.bot.send_message(
            user.telegram_id,
            f"❌ Ваша бронь на {format_date_ru(trip.date)} отменена администратором",
        )
    except Exception:
        pass

    for admin_id in ADMIN_IDS:
        if admin_id == callback.from_user.id:
            continue
        try:
            await callback.bot.send_message(
                admin_id,
                f"❌ Бронь #{booking_id} отменена администратором\n"
                f"👤 Пользователь: @{user.username or user.telegram_id}\n"
                f"🚌 Рейс: {format_date_ru(trip.date)}",
            )
        except Exception:
            pass


@router.callback_query(F.data == "admin_create_trip")
async def admin_create_trip_start(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    await state.set_state(AdminStates.create_date)
    await callback.message.edit_text(
        "📦 Введите дату рейса в формате ДД.ММ.ГГГГ (например, 15.08.2026):",
        reply_markup=back_button("admin_menu"),
    )


@router.message(AdminStates.create_date)
async def admin_create_trip_date(message: Message, state: FSMContext):
    try:
        day, month, year = map(int, message.text.strip().split("."))
        trip_date = date(year, month, day)
    except (ValueError, IndexError):
        await message.answer("❌ Неверный формат\nИспользуйте ДД.ММ.ГГГГ (например, 15.08.2026):")
        return

    await state.update_data(trip_date=trip_date)
    await state.set_state(AdminStates.create_bus_type)
    await message.answer(
        "🚌 Выберите тип автобуса:",
        reply_markup=bus_type_kb(),
    )


@router.callback_query(AdminStates.create_bus_type, F.data.in_(["bus_type_vip", "bus_type_sleeper"]))
async def admin_create_trip_bus_type(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    bus_type = "sleeper" if callback.data == "bus_type_sleeper" else "vip"
    await state.update_data(bus_type=bus_type)
    await state.set_state(AdminStates.create_seats_left)
    await callback.message.edit_text(
        "📦 Введите количество мест в левом ряду (например, 20):",
        reply_markup=back_button("admin_menu"),
    )


@router.message(AdminStates.create_seats_left)
async def admin_create_trip_seats_left(message: Message, state: FSMContext):
    if not message.text.strip().isdigit():
        await message.answer("❌ Введите целое число")
        return
    seats = int(message.text.strip())
    if seats < 1:
        await message.answer("❌ Количество мест должно быть больше 0")
        return

    data = await state.get_data()
    await state.update_data(seats_left=seats)

    if data.get("bus_type") == "sleeper":
        await state.set_state(AdminStates.create_seats_middle)
        await message.answer(
            "📦 Введите количество мест в среднем ряду (например, 20):",
            reply_markup=back_button("admin_menu"),
        )
    else:
        await state.update_data(seats_middle=0)
        await state.set_state(AdminStates.create_seats_right)
        await message.answer(
            "📦 Введите количество мест в правом ряду (например, 20):",
            reply_markup=back_button("admin_menu"),
        )


@router.message(AdminStates.create_seats_middle)
async def admin_create_trip_seats_middle(message: Message, state: FSMContext):
    if not message.text.strip().isdigit():
        await message.answer("❌ Введите целое число")
        return
    seats = int(message.text.strip())
    if seats < 1:
        await message.answer("❌ Количество мест должно быть больше 0")
        return

    await state.update_data(seats_middle=seats)
    await state.set_state(AdminStates.create_seats_right)
    await message.answer(
        "📦 Введите количество мест в правом ряду (например, 20):",
        reply_markup=back_button("admin_menu"),
    )


@router.message(AdminStates.create_seats_right)
async def admin_create_trip_seats_right(message: Message, state: FSMContext):
    if not message.text.strip().isdigit():
        await message.answer("❌ Введите целое число")
        return
    seats = int(message.text.strip())
    if seats < 1:
        await message.answer("❌ Количество мест должно быть больше 0")
        return

    await state.update_data(seats_right=seats)
    await state.set_state(AdminStates.create_price_adult)
    await message.answer("💰 Введите цену для взрослого (в VND, например, 1450000):")


@router.message(AdminStates.create_price_adult)
async def admin_create_trip_price_adult(message: Message, state: FSMContext):
    if not message.text.strip().isdigit():
        await message.answer("❌ Введите целое число")
        return
    price = int(message.text.strip())
    if price < 1:
        await message.answer("❌ Цена должна быть больше 0")
        return

    await state.update_data(price_adult=price)
    await state.set_state(AdminStates.create_price_child)
    await message.answer("💰 Введите цену для ребёнка (в VND, например, 1000000):")


@router.message(AdminStates.create_price_child)
async def admin_create_trip_price_child(message: Message, state: FSMContext):
    if not message.text.strip().isdigit():
        await message.answer("❌ Введите целое число")
        return
    price = int(message.text.strip())
    if price < 1:
        await message.answer("❌ Цена должна быть больше 0")
        return

    await state.update_data(price_child=price)
    await state.set_state(AdminStates.create_location)
    await message.answer("📍 Введите место сбора (адрес в Нячанге):")


@router.message(AdminStates.create_location)
async def admin_create_trip_location(message: Message, state: FSMContext):
    location = message.text.strip()
    if not location:
        await message.answer("❌ Место сбора не может быть пустым")
        return

    await state.update_data(location=location)
    await state.set_state(AdminStates.create_time)
    await message.answer(
        "⏰ Введите время отправления в формате ЧЧ:ММ (например, 06:00):"
    )


@router.message(AdminStates.create_time)
async def admin_create_trip_time(message: Message, state: FSMContext):
    time_str = message.text.strip()
    import re
    if not re.match(r"^\d{2}:\d{2}$", time_str):
        await message.answer("❌ Неверный формат\nИспользуйте ЧЧ:ММ (например, 06:00):")
        return
    try:
        h, m = map(int, time_str.split(":"))
        if h < 0 or h > 23 or m < 0 or m > 59:
            raise ValueError
    except ValueError:
        await message.answer("❌ Неверное время\nЧасы 0-23, минуты 0-59:")
        return
    await state.update_data(departure_time=time_str)
    await state.set_state(AdminStates.create_bus)
    await message.answer(
        "🚌 Введите номер автобуса (или отправьте «-», чтобы пропустить):"
    )


@router.message(AdminStates.create_bus)
async def admin_create_trip_bus(message: Message, state: FSMContext):
    bus = message.text.strip()
    if bus == "-":
        bus = ""

    data = await state.get_data()
    trip_date = data["trip_date"]
    bus_type = data["bus_type"]
    seats_left = data["seats_left"]
    seats_middle = data.get("seats_middle", 0)
    seats_right = data["seats_right"]
    price_adult = data["price_adult"]
    price_child = data["price_child"]
    location = data["location"]
    departure_time = data.get("departure_time")
    total_seats = seats_left + seats_middle + seats_right

    async with async_session_maker() as session:
        trip = Trip(
            date=trip_date,
            max_seats=total_seats,
            bus_type=bus_type,
            seats_left=seats_left,
            seats_middle=seats_middle,
            seats_right=seats_right,
            price_adult=price_adult,
            price_child=price_child,
            pickup_location=location,
            departure_time=departure_time,
            bus_number=bus,
            is_active=True,
            created_at=now_utc(),
        )
        session.add(trip)
        await session.commit()
        await session.refresh(trip)

    await state.clear()

    bus_type_label = "VIP" if bus_type == "vip" else "Sleeper"
    seat_layout = f"{seats_left}L"
    if bus_type == "sleeper":
        seat_layout += f" + {seats_middle}M"
    seat_layout += f" + {seats_right}R"

    await message.answer(
        f"✅ Рейс создан!\n\n"
        f"📅 Дата: {format_date_ru(trip_date)}\n"
        f"⏰ Отправление: {departure_time or '—'}\n"
        f"🚌 Тип: {bus_type_label}\n"
        f"💺 Схема: {seat_layout} ({total_seats} мест)\n"
        f"💰 Цена взр.: {format_price(price_adult)}\n"
        f"💰 Цена дет.: {format_price(price_child)}\n"
        f"📍 Сбор: {location}\n"
        f"🚌 Номер: {bus or '—'}",
        reply_markup=admin_main_menu(),
    )


@router.callback_query(F.data.startswith("admin_edit_"))
async def admin_edit_trip_menu(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    parts = callback.data.split("_")
    trip_id = int(parts[2])

    if len(parts) == 3:
        async with async_session_maker() as session:
            result = await session.execute(select(Trip).where(Trip.id == trip_id))
            trip = result.scalar_one_or_none()

        if not trip:
            await callback.message.edit_text("🚌 Рейс не найден", reply_markup=back_button("admin_trips"))
            return

        await callback.message.edit_text(
            "✏️ Что вы хотите изменить?",
            reply_markup=admin_edit_trip(trip_id),
        )
        return

    field = parts[3]

    async with async_session_maker() as session:
        result = await session.execute(select(Trip).where(Trip.id == trip_id))
        trip = result.scalar_one_or_none()

    if not trip:
        await callback.message.edit_text("🚌 Рейс не найден", reply_markup=back_button("admin_trips"))
        return

    if field == "bus_type":
        await state.update_data(edit_trip_id=trip_id, edit_field=field)
        await state.set_state(AdminStates.edit_value)
        await callback.message.edit_text(
            f"ℹ️ Текущий тип: {'VIP' if trip.bus_type == 'vip' else 'Sleeper'}\n"
            f"✏️ Выберите новый тип:",
            reply_markup=bus_type_kb(),
        )
        return

    field_names = {
        "date": "📅 новую дату (ДД.ММ.ГГГГ)",
        "seats_left": "💺 новое количество мест слева",
        "seats_middle": "💺 новое количество мест в центре",
        "seats_right": "💺 новое количество мест справа",
        "price_adult": "💰 новую цену для взрослого",
        "price_child": "💰 новую цену для ребёнка",
        "location": "📍 новое место сбора",
        "time": "⏰ новое время отправления (ЧЧ:ММ)",
        "bus": "🚌 новый номер автобуса",
    }

    current_values = {
        "date": format_date_ru(trip.date),
        "seats_left": str(trip.seats_left),
        "seats_middle": str(trip.seats_middle),
        "seats_right": str(trip.seats_right),
        "price_adult": format_price(trip.price_adult),
        "price_child": format_price(trip.price_child),
        "location": trip.pickup_location,
        "time": trip.departure_time or "-",
        "bus": trip.bus_number or "-",
    }

    await state.update_data(edit_trip_id=trip_id, edit_field=field)
    await state.set_state(AdminStates.edit_value)

    await callback.message.edit_text(
        f"ℹ️ Текущее значение: {current_values.get(field, '?')}\n"
        f"✏️ Введите {field_names.get(field, 'ℹ️ новое значение')}:",
        reply_markup=back_button(f"admin_edit_{trip_id}"),
    )


@router.message(AdminStates.edit_value)
async def admin_process_edit(message: Message, state: FSMContext):
    data = await state.get_data()
    trip_id = data.get("edit_trip_id")
    field = data.get("edit_field")

    if not trip_id or not field:
        await message.answer("❌ Ошибка\nПопробуйте снова", reply_markup=admin_main_menu())
        await state.clear()
        return

    value = message.text.strip()

    async with async_session_maker() as session:
        result = await session.execute(select(Trip).where(Trip.id == trip_id))
        trip = result.scalar_one_or_none()

        if not trip:
            await message.answer("🚌 Рейс не найден", reply_markup=admin_main_menu())
            await state.clear()
            return

        try:
            if field == "date":
                day, month, year = map(int, value.split("."))
                trip.date = date(year, month, day)
            elif field == "seats_left":
                seats = int(value)
                if seats < 0:
                    raise ValueError
                trip.seats_left = seats
                trip.max_seats = get_trip_total_seats(trip)
            elif field == "seats_middle":
                seats = int(value)
                if seats < 0:
                    raise ValueError
                trip.seats_middle = seats
                trip.max_seats = get_trip_total_seats(trip)
            elif field == "seats_right":
                seats = int(value)
                if seats < 0:
                    raise ValueError
                trip.seats_right = seats
                trip.max_seats = get_trip_total_seats(trip)
            elif field == "price_adult":
                price = int(value)
                if price < 1:
                    raise ValueError
                trip.price_adult = price
            elif field == "price_child":
                price = int(value)
                if price < 1:
                    raise ValueError
                trip.price_child = price
            elif field == "location":
                if not value:
                    raise ValueError
                trip.pickup_location = value
            elif field == "time":
                import re
                if not re.match(r"^\d{2}:\d{2}$", value):
                    await message.answer("❌ Неверный формат\nИспользуйте ЧЧ:ММ (например, 06:00):")
                    return
                h, m = map(int, value.split(":"))
                if h < 0 or h > 23 or m < 0 or m > 59:
                    raise ValueError
                trip.departure_time = value if value != "-" else None
            elif field == "bus":
                trip.bus_number = value if value != "-" else ""
            else:
                await message.answer("❌ Неизвестное поле", reply_markup=admin_main_menu())
                await state.clear()
                return

            await session.commit()
        except (ValueError, IndexError):
            await message.answer("❌ Неверный формат значения\nПопробуйте снова")
            return

    await state.clear()

    msg = "✅ Рейс обновлён!"
    if field in ("seats_left", "seats_middle", "seats_right"):
        async with async_session_maker() as session:
            occupied = await get_occupied_seats(session, trip_id)
        new_total = get_trip_total_seats(trip)
        shortage = occupied - new_total
        if shortage > 0:
            msg += (
                f"\n\n⚠️ Внимание! Теперь всего {new_total} мест, "
                f"но занято {occupied} (❌ не хватает {shortage} мест).\n"
                f"📋 Используйте «Оповестить всех», чтобы сообщить пассажирам."
            )
        elif new_total > occupied:
            msg += f"\n💚 Свободных мест после изменения: {new_total - occupied}"

    await message.answer(
        msg,
        reply_markup=back_button(f"admin_trip_{trip_id}"),
    )


@router.callback_query(AdminStates.edit_value, F.data.in_(["bus_type_vip", "bus_type_sleeper"]))
async def admin_edit_bus_type(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    data = await state.get_data()
    trip_id = data.get("edit_trip_id")
    field = data.get("edit_field")

    if field != "bus_type" or not trip_id:
        await state.clear()
        return

    new_type = "sleeper" if callback.data == "bus_type_sleeper" else "vip"

    async with async_session_maker() as session:
        result = await session.execute(select(Trip).where(Trip.id == trip_id))
        trip = result.scalar_one_or_none()
        if trip and trip.bus_type == "vip" and new_type == "sleeper":
            trip.seats_middle = trip.seats_middle or trip.seats_left // 2
        elif trip and trip.bus_type == "sleeper" and new_type == "vip":
            trip.seats_middle = 0
        if trip:
            trip.bus_type = new_type
            trip.max_seats = get_trip_total_seats(trip)
            await session.commit()

    await state.clear()
    if not trip:
        await callback.message.edit_text("🚌 Рейс не найден", reply_markup=admin_main_menu())
        return

    async with async_session_maker() as session:
        occupied = await get_occupied_seats(session, trip_id)
    new_total = get_trip_total_seats(trip)
    shortage = occupied - new_total

    msg = "✅ Тип автобуса обновлён!"
    if shortage > 0:
        msg += (
            f"\n\n⚠️ Внимание! Новый автобус вмещает только {new_total} мест, "
            f"но уже занято {occupied} (❌ не хватает {shortage} мест).\n"
            f"📋 Используйте «Оповестить всех», чтобы сообщить пассажирам об изменениях."
        )
    elif new_total > occupied:
        msg += f"\n💚 Свободных мест после изменения: {new_total - occupied}"

    await callback.message.edit_text(
        msg,
        reply_markup=back_button(f"admin_trip_{trip_id}"),
    )


@router.callback_query(F.data.startswith("admin_delete_"))
async def admin_delete_trip(callback: CallbackQuery):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    parts = callback.data.split("_")
    if "confirm" in parts:
        trip_id = int(parts[-1])
        async with async_session_maker() as session:
            result = await session.execute(select(Trip).where(Trip.id == trip_id))
            trip = result.scalar_one_or_none()
            if not trip:
                await callback.message.edit_text("🚌 Рейс не найден")
                return

            booking_result = await session.execute(
                select(Booking).where(Booking.trip_id == trip_id)
            )
            bookings = booking_result.scalars().all()
            date_str = format_date_ru(trip.date)
            user_notify = [(b.id, b.user.telegram_id) for b in bookings]
            paid_count = sum(1 for b in bookings if b.status == "paid")
            total_refund = sum(b.total_price + (b.pending_add_amount or 0) for b in bookings if b.status == "paid")

            for b in bookings:
                await session.delete(b)
            await session.delete(trip)
            await session.commit()

        for bid, uid in user_notify:
            try:
                await callback.bot.send_message(
                    uid,
                    f"❌ Рейс на {date_str} отменён администратором.\n"
                    f"Ваша бронь #{bid} отменена.",
                )
            except Exception:
                pass

        result_lines = [f"🗑 Рейс на {date_str} удалён.", f"📋 Отменено броней: {len(user_notify)}"]
        if paid_count:
            result_lines.append(f"✅ Из них оплачено: {paid_count}")
            result_lines.append(f"💰 К возврату: {format_price(total_refund)}")
        await callback.message.edit_text(
            "\n".join(result_lines),
            reply_markup=back_button("admin_trips"),
        )
        return

    trip_id = int(parts[-1])
    async with async_session_maker() as session:
        result = await session.execute(select(Trip).where(Trip.id == trip_id))
        trip = result.scalar_one_or_none()

        if not trip:
            await callback.message.edit_text("🚌 Рейс не найден")
            return

        booking_result = await session.execute(
            select(Booking).where(Booking.trip_id == trip_id)
        )
        bookings = booking_result.scalars().all()

    total_bookings = len(bookings)
    paid_bookings = [b for b in bookings if b.status == "paid"]
    total_refund = sum(b.total_price + (b.pending_add_amount or 0) for b in paid_bookings)

    lines = [
        f"⚠️ Вы уверены, что хотите удалить рейс на {format_date_ru(trip.date)}?",
        f"📋 Всего броней: {total_bookings}",
    ]

    if paid_bookings:
        lines.append("")
        lines.append("✅ ОПЛАЧЕНО (требуется возврат):")
        for b in paid_bookings:
            full_name = b.user.full_name or b.user.username or str(b.user.telegram_id)
            total = b.total_price + (b.pending_add_amount or 0)
            lines.append(f"  • {full_name} — {format_price(total)}")
        lines.append(f"  💰 Итого к возврату: {format_price(total_refund)}")

    lines.append("")
    lines.append("Пользователи получат уведомление об отмене.")

    await callback.message.edit_text(
        "\n".join(lines),
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="✅ Да, удалить", callback_data=f"admin_delete_confirm_{trip_id}")],
                [InlineKeyboardButton(text="❌ Нет", callback_data=f"admin_trip_{trip_id}")],
            ]
        ),
    )


@router.callback_query(F.data.startswith("admin_stats_"))
async def admin_show_stats(callback: CallbackQuery):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    trip_id = int(callback.data.split("_")[-1])
    async with async_session_maker() as session:
        stats = await get_trip_stats(session, trip_id)

    if not stats:
        await callback.message.edit_text("🚌 Рейс не найден", reply_markup=back_button("admin_trips"))
        return

    trip = stats["trip"]
    total_seats = get_trip_total_seats(trip)
    text = (
        f"📊 Статистика рейса: {format_date_ru(trip.date)}\n\n"
        f"💺 Всего мест: {total_seats}\n"
        f"👥 Занято: {stats['occupied']}\n"
        f"  ✅ Оплачено: {stats['paid']}\n"
        f"  ⏳ Ожидает оплаты: {stats['pending_seats']}\n"
        f"💚 Свободно: {stats['free']}"
    )

    await callback.message.edit_text(text, reply_markup=back_button(f"admin_trip_{trip_id}"))


@router.callback_query(F.data.startswith("admin_export_"))
async def admin_export_trip(callback: CallbackQuery):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    trip_id = int(callback.data.split("_")[-1])

    async with async_session_maker() as session:
        result = await session.execute(select(Trip).where(Trip.id == trip_id))
        trip = result.scalar_one_or_none()

        if not trip:
            await callback.message.edit_text("🚌 Рейс не найден")
            return

        booking_result = await session.execute(
            select(Booking).where(
                and_(Booking.trip_id == trip_id, Booking.status == "paid")
            ).order_by(Booking.created_at)
        )
        bookings = booking_result.scalars().all()

    lines = [
        f"🚌 РЕЙС: {format_date_ru(trip.date)}",
        f"🚌 Тип: {'VIP' if trip.bus_type == 'vip' else 'Sleeper'}",
        f"🚌 Номер: {trip.bus_number or '—'}",
        f"📍 Сбор: {trip.pickup_location}",
        "",
        "📋 Список участников (оплачено):",
        "",
    ]

    if not bookings:
        lines.append("📋 Нет оплаченных броней")
    else:
        lines.append(f"{'N':<4}{'ID':<6}{'Пассажир':<20}{'Взр':<5}{'Дет':<5}{'Места':<12}{'Статус':<10}")
        lines.append("-" * 60)
        for i, b in enumerate(bookings, 1):
            username = b.user.username or str(b.user.telegram_id)
            status = STATUS_EMOJI.get(b.status, b.status)
            seats_str = b.selected_seats or "—"
            lines.append(
                f"{i:<4}#{b.id:<4}@{username:<18}{b.adults:<5}{b.children:<5}{seats_str:<12}{status:<10}"
            )

    total_people = sum(b.adults + b.children for b in bookings)
    lines.append("")
    lines.append(f"👥 Всего оплачено: {total_people} человек(а)")

    await callback.message.edit_text(
        "\n".join(lines),
        reply_markup=back_button(f"admin_trip_{trip_id}"),
    )


@router.callback_query(F.data.startswith("admin_export_xlsx_"))
async def admin_export_xlsx(callback: CallbackQuery):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    trip_id = int(callback.data.split("_")[-1])

    async with async_session_maker() as session:
        result = await session.execute(select(Trip).where(Trip.id == trip_id))
        trip = result.scalar_one_or_none()

        if not trip:
            await callback.message.edit_text("🚌 Рейс не найден")
            return

        booking_result = await session.execute(
            select(Booking).where(Booking.trip_id == trip_id).order_by(Booking.created_at)
        )
        bookings = booking_result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{trip.date}"

    header_fill = PatternFill(start_color="007AFF", end_color="007AFF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    pending_fill = PatternFill(start_color="FFD5D5", end_color="FFD5D5", fill_type="solid")

    headers = ["№", "ID брони", "ФИО", "Username", "Взрослые", "Дети", "Всего", "Места", "Сумма", "Способ оплаты", "Статус", "Долг", "Комментарий к переводу", "История платежей"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    debt_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")

    for i, b in enumerate(bookings, 1):
        row = i + 1
        full_name = b.user.full_name or "—"
        username = f"@{b.user.username}" if b.user.username else str(b.user.telegram_id)
        total_people = b.adults + b.children
        status_label = STATUS_EMOJI.get(b.status, b.status)
        payment_label = {
            "cash_vnd": "Наличные (VND)",
            "transfer_rub": "Перевод (RUB)",
            "transfer_kzt": "Перевод (KZT)",
        }.get(b.payment_method, b.payment_method or "—")
        if b.pending_add_amount:
            debt = format_price(b.pending_add_amount)
            if b.pending_add_people:
                debt = f"+{b.pending_add_people}чел · {debt}"
        else:
            debt = "—"
        if b.payment_method and b.payment_method != "cash_vnd":
            payment_comment = (b.payment_comment or "").strip() or full_name
        else:
            payment_comment = "—"

        history_lines = []
        for p in b.payments:
            mark = "✅" if p.status == "paid" else "⏳"
            cmt = (p.comment or "").strip()
            history_lines.append(
                f"{mark} {payment_method_label(p.method)} {format_price(p.amount)}"
                + (f" ({cmt})" if cmt else "")
            )
        history_str = "\n".join(history_lines)

        values = [
            i, f"#{b.id}", full_name, username,
            b.adults, b.children, total_people,
            b.selected_seats or "—",
            format_price(b.total_price), payment_label, status_label, debt, payment_comment,
            history_str,
        ]

        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if b.status == "pending":
                cell.fill = pending_fill
            elif b.pending_add_amount:
                cell.fill = debt_fill

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 26
    ws.column_dimensions["M"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    await callback.message.edit_text(
        "📄 Excel файл готов, отправляю...",
        reply_markup=back_button(f"admin_trip_{trip_id}"),
    )
    await callback.bot.send_document(
        callback.from_user.id,
        BufferedInputFile(buf.getvalue(), filename=f"рейс_{trip.date}.xlsx"),
        caption=f"📄 Экспорт броней на {format_date_ru(trip.date)}",
    )


@router.callback_query(F.data.startswith("admin_broadcast_"))
async def admin_broadcast_start(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return
    trip_id = int(callback.data.split("_")[-1])

    async with async_session_maker() as session:
        result = await session.execute(select(Trip).where(Trip.id == trip_id))
        trip = result.scalar_one_or_none()

    if not trip:
        await callback.message.edit_text("🚌 Рейс не найден", reply_markup=admin_main_menu())
        return

    await state.update_data(broadcast_trip_id=trip_id, trip_date=format_date_ru(trip.date))
    await state.set_state(AdminStates.broadcast_text)
    await callback.message.edit_text(
        f"📢 Напишите сообщение для всех пассажиров рейса {format_date_ru(trip.date)}:\n\n"
        f"ℹ️ Сообщение будет отправлено всем с активной бронью (ожидает оплаты / оплачено).\n"
        f"❌ Отправьте «-» чтобы отменить.",
        reply_markup=back_button(f"admin_trip_{trip_id}"),
    )


@router.message(AdminStates.broadcast_text)
async def admin_broadcast_message(message: Message, state: FSMContext):
    text = message.text.strip()
    if text == "-":
        await state.clear()
        await message.answer("❌ Рассылка отменена", reply_markup=admin_main_menu())
        return

    if len(text) < 2:
        await message.answer("❌ Сообщение слишком короткое. Напишите текст или отправьте «-» для отмены:")
        return

    data = await state.get_data()
    trip_id = data["broadcast_trip_id"]
    trip_date = data["trip_date"]

    async with async_session_maker() as session:
        result = await session.execute(
            select(Booking).where(
                and_(Booking.trip_id == trip_id, Booking.status.in_(["pending", "paid"]))
            )
        )
        bookings = result.scalars().all()
        user_ids = list({b.user.telegram_id for b in bookings if b.user})

    await state.update_data(broadcast_text=text)
    await message.answer(
        f"📢 Предпросмотр рассылки:\n\n"
        f"🚌 Рейс: {trip_date}\n"
        f"📨 Получателей: {len(user_ids)}\n"
        f"────────────────\n"
        f"{text}\n"
        f"────────────────",
        reply_markup=broadcast_confirm_kb(),
    )


@router.callback_query(F.data == "broadcast_send")
async def admin_broadcast_send(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    data = await state.get_data()
    trip_id = data.get("broadcast_trip_id")
    text = data.get("broadcast_text")

    if not trip_id or not text:
        await callback.message.edit_text("❌ Ошибка данных", reply_markup=admin_main_menu())
        await state.clear()
        return

    async with async_session_maker() as session:
        result = await session.execute(
            select(Booking).where(
                and_(Booking.trip_id == trip_id, Booking.status.in_(["pending", "paid"]))
            )
        )
        bookings = result.scalars().all()
        user_ids = list({b.user.telegram_id for b in bookings if b.user})

    trip_date = data.get("trip_date", "?")

    sent = 0
    failed = 0
    for uid in user_ids:
        try:
            await callback.bot.send_message(
                uid,
                f"📢 Уведомление по рейсу {trip_date}\n\n{text}",
            )
            sent += 1
        except Exception:
            failed += 1

    await state.clear()
    await callback.message.edit_text(
        f"✅ Рассылка завершена!\n\n"
        f"📨 Отправлено: {sent}\n"
        f"❌ Ошибок: {failed}",
        reply_markup=back_button(f"admin_trip_{trip_id}"),
    )


@router.callback_query(F.data == "broadcast_cancel")
async def admin_broadcast_cancel(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    data = await state.get_data()
    trip_id = data.get("broadcast_trip_id")
    await state.clear()
    await callback.message.edit_text(
        "❌ Рассылка отменена",
        reply_markup=back_button(f"admin_trip_{trip_id}") if trip_id else admin_main_menu(),
    )
